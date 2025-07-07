from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, make_response
from datetime import datetime, timedelta
from collections import defaultdict
from decimal import Decimal
from data import get_db_connection
import pandas as pd
import io
import locale
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

productos_bp = Blueprint('productos', __name__, url_prefix='/productos')

@productos_bp.route('/', methods=['GET', 'POST'])
def mostrar_y_agregar_productos():
    if request.method == 'POST':
        try:
            nombre = request.form['nombre'].strip()
            descripcion = request.form.get('descripcion', '').strip()
            precio = float(request.form['precio'])
            stock = int(request.form['stock'])
            proveedor_id = int(request.form['proveedor_id'])
            categoria_id = int(request.form['categoria_id'])
            fecha_registro = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            if not nombre:
                flash('El nombre del producto es obligatorio.', 'danger')
            elif precio < 0 or stock < 0:
                flash('Precio y stock deben ser valores positivos.', 'danger')
            else:
                conn = get_db_connection()
                cursor = conn.cursor()

                # Verificar si ya existe un producto con el mismo nombre, proveedor y categoría
                cursor.execute("""
                    SELECT id, stock FROM productos
                    WHERE LOWER(nombre) = %s AND proveedor_id = %s AND categoria_id = %s
                """, (nombre.lower(), proveedor_id, categoria_id))
                existente = cursor.fetchone()

                if existente:
                    # Actualizar stock y otros datos si quieres
                    nuevo_stock = existente['stock'] + stock
                    cursor.execute("""
                        UPDATE productos
                        SET stock = %s, precio = %s, descripcion = %s
                        WHERE id = %s
                    """, (nuevo_stock, precio, descripcion, existente['id']))
                    # ⚠️ Aquí el flash con categoría especial
                    flash(f'Se encontró un producto existente y se actualizó el stock a {nuevo_stock}.', 'stock_similar')
                else:
                    cursor.execute("""
                        INSERT INTO productos(nombre, descripcion, precio, stock, proveedor_id, fecha_registro, categoria_id)
                        VALUES (%s, %s, %s, %s, %s, %s, %s)
                    """, (nombre, descripcion, precio, stock, proveedor_id, fecha_registro, categoria_id))
                    flash('✅ Producto agregado correctamente.', 'success')

                conn.commit()
                cursor.close()
                conn.close()
                return redirect(url_for('productos.mostrar_y_agregar_productos'))

        except Exception as e:
            flash(f'❌ Error al agregar producto: {e}', 'danger')

    # Mostrar formulario y tabla
    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT p.*, c.nombre AS categoria_nombre
        FROM productos p
        LEFT JOIN categoria c ON p.categoria_id = c.id
    """)
    productos = cursor.fetchall()

    cursor.execute("SELECT id, nombre, documento_tipo, documento_numero FROM proveedores")
    proveedores = cursor.fetchall()

    cursor.execute("SELECT id, nombre FROM categoria WHERE estado = 1")
    categorias = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template(
        'productos/productos.html',
        productos=productos,
        producto=None,
        proveedores=proveedores,
        categorias=categorias
    )

@productos_bp.route('/editar/<int:id>', methods=['GET', 'POST'])
def editar_producto(id):
    conn = get_db_connection()
    cursor = conn.cursor()

    if request.method == 'POST':
        try:
            nombre = request.form['nombre'].strip()
            descripcion = request.form.get('descripcion', '').strip()
            precio = float(request.form['precio'])
            stock = int(request.form['stock'])
            proveedor_id = int(request.form['proveedor_id'])
            categoria_id = int(request.form['categoria_id'])

            if not nombre:
                flash('El nombre del producto es obligatorio.', 'danger')
            elif precio < 0 or stock < 0:
                flash('Precio y stock deben ser valores positivos.', 'danger')
            else:
                cursor.execute("""
                    UPDATE productos 
                    SET nombre=%s, descripcion=%s, precio=%s, stock=%s, proveedor_id=%s, categoria_id=%s
                    WHERE id=%s
                """, (nombre, descripcion, precio, stock, proveedor_id, categoria_id, id))
                conn.commit()
                flash(f'Producto "{nombre}" actualizado correctamente.', 'producto_actualizado')
        except Exception as e:
            flash(f'Error al actualizar producto: {e}', 'danger')

    # Obtener datos del producto a editar
    cursor.execute("""
        SELECT p.*, c.nombre AS categoria_nombre
        FROM productos p
        LEFT JOIN categoria c ON p.categoria_id = c.id
        WHERE p.id = %s
    """, (id,))
    producto = cursor.fetchone()

    # Obtener lista de productos
    cursor.execute("""
        SELECT p.*, c.nombre AS categoria_nombre
        FROM productos p
        LEFT JOIN categoria c ON p.categoria_id = c.id
    """)
    productos = cursor.fetchall()

    # Obtener proveedores
    cursor.execute("SELECT id, nombre, documento_tipo, documento_numero FROM proveedores")
    proveedores = cursor.fetchall()

    # Obtener categorías activas
    cursor.execute("SELECT id, nombre FROM categoria WHERE estado = 1")
    categorias = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template(
        'productos/productos.html',
        productos=productos,
        producto=producto,
        proveedores=proveedores,
        categorias=categorias
    )

@productos_bp.route('/eliminar/<int:id>', methods=['POST'])
def eliminar_producto(id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM productos WHERE id = %s", (id,))
        conn.commit()
        cursor.close()
        conn.close()
        flash(f'Producto {id} eliminado correctamente', 'success')
    except Exception as e:
        flash(f'Error al eliminar producto: {e}', 'danger')
    return redirect(url_for('productos.mostrar_y_agregar_productos'))

@productos_bp.route('/stock')
def stock():
    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT 
            p.id, p.codigo_producto, p.nombre, p.descripcion, p.precio, p.stock, 
            c.nombre AS categoria, 
            pr.nombre AS proveedor_nombre, 
            pr.documento_tipo, pr.documento_numero
        FROM productos p
        JOIN proveedores pr ON p.proveedor_id = pr.id
        JOIN categoria c ON p.categoria_id = c.id
        GROUP BY p.id, p.codigo_producto, p.nombre, p.descripcion, p.precio, p.stock, c.nombre, proveedor_nombre, pr.documento_tipo, pr.documento_numero
    """)
    productos = cursor.fetchall()

    cursor.execute("SELECT id, nombre FROM categoria ORDER BY nombre")
    categorias_raw = cursor.fetchall()
    categorias = [row['nombre'] for row in categorias_raw]

    cursor.close()
    conn.close()

    # Alerta si hay productos con bajo o sin stock
    umbral_stock_bajo = 10
    productos_sin_stock = [p for p in productos if p['stock'] == 0]
    productos_bajo_stock = [p for p in productos if 0 < p['stock'] <= umbral_stock_bajo]

    if productos_sin_stock or productos_bajo_stock:
        mensaje = '⚠ Alerta: '
        if productos_sin_stock:
            mensaje += f'{len(productos_sin_stock)} producto(s) están agotados. '
        if productos_bajo_stock:
            mensaje += f'{len(productos_bajo_stock)} producto(s) tienen bajo stock (≤ {umbral_stock_bajo}).'
        flash(mensaje.strip(), 'warning')

    return render_template('productos/stock.html', productos=productos, categorias=categorias, categoria_seleccionada=None, nombre_busqueda='', codigo_producto_busqueda='')


@productos_bp.route('/buscar_stock', methods=['POST'])
def buscar_stock():
    categoria = request.form.get('categoria', '').strip()
    nombre = request.form.get('nombre', '').strip()
    codigo_producto = request.form.get('codigo_producto', '').strip()

    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        sql = """
            SELECT 
                p.id, p.codigo_producto, p.nombre, p.descripcion, p.precio, p.stock, 
                c.nombre AS categoria,
                pr.nombre AS proveedor_nombre, 
                pr.documento_tipo, pr.documento_numero
            FROM productos p
            JOIN proveedores pr ON p.proveedor_id = pr.id
            JOIN categoria c ON p.categoria_id = c.id
            WHERE 1=1
        """
        params = []

        if categoria and categoria.lower() != 'todas':
            sql += " AND c.nombre = %s"
            params.append(categoria)

        if nombre:
            sql += " AND LOWER(p.nombre) LIKE %s"
            params.append(f"%{nombre.lower()}%")

        if codigo_producto:
            sql += " AND p.codigo_producto LIKE %s"
            params.append(f"%{codigo_producto}%")

        sql += """
            GROUP BY p.id, p.codigo_producto, p.nombre, p.descripcion, p.precio, p.stock, categoria, proveedor_nombre, pr.documento_tipo, pr.documento_numero
        """

        cursor.execute(sql, tuple(params))
        productos = cursor.fetchall()

        cursor.execute("SELECT nombre FROM categoria ORDER BY nombre ASC")
        categorias_raw = cursor.fetchall()
        categorias = [row['nombre'] for row in categorias_raw]

        cursor.close()
        conn.close()

        # Alerta si hay productos con poco o sin stock
        umbral_stock_bajo = 10
        productos_sin_stock = [p for p in productos if p['stock'] == 0]
        productos_bajo_stock = [p for p in productos if 0 < p['stock'] <= umbral_stock_bajo]

        if productos_sin_stock or productos_bajo_stock:
            mensaje = '⚠ Alerta: '
            if productos_sin_stock:
                mensaje += f'{len(productos_sin_stock)} producto(s) están agotados. '
            if productos_bajo_stock:
                mensaje += f'{len(productos_bajo_stock)} producto(s) tienen bajo stock (≤ {umbral_stock_bajo}).'
            flash(mensaje.strip(), 'warning')

        return render_template(
            'productos/stock.html',
            productos=productos,
            categorias=categorias,
            categoria_seleccionada=categoria,
            nombre_busqueda=nombre,
            codigo_producto_busqueda=codigo_producto
        )

    except Exception as e:
        flash(f'Error al buscar productos: {e}', 'danger')
        return redirect(url_for('productos.stock'))


@productos_bp.route('/actualizar_stock/<int:id>', methods=['POST'])
def actualizar_stock(id):
    try:
        nuevo_stock = int(request.form['nuevo_stock'])
        if nuevo_stock < 0:
            flash('El stock no puede ser negativo.', 'danger')
        else:
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute("UPDATE productos SET stock = %s WHERE id = %s", (nuevo_stock, id))
            conn.commit()
            cursor.close()
            conn.close()
            flash(f'Stock del producto {id} actualizado a {nuevo_stock}.', 'success')
    except Exception as e:
        flash(f'Error al actualizar stock: {e}', 'danger')
    return redirect(url_for('productos.stock'))


@productos_bp.route('/vendidos')
def productos_vendidos():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        numero = request.args.get('numero', '').strip().lower()
        fecha = request.args.get('fecha', '').strip()
        tipo = request.args.get('tipo', '').strip().lower()

        cursor.execute("""
            SELECT 
                f.id AS factura_id,
                f.numero_factura,
                f.tipo_documento,
                f.fecha_emision,
                p.nombre AS producto_nombre,
                df.cantidad,
                df.precio_unitario,
                (df.cantidad * df.precio_unitario) AS subtotal,
                c.documento_tipo AS cliente_tipo_documento,
                c.documento_numero AS cliente_numero_documento,
                c.nombre AS cliente_nombre
            FROM detalle_factura df
            JOIN productos p ON df.producto_id = p.id
            JOIN facturas f ON df.factura_id = f.id
            JOIN clientes c ON f.cliente_documento_numero = c.documento_numero
            ORDER BY f.fecha_emision DESC, f.numero_factura;
        """)

        filas = cursor.fetchall()

        comprobantes = defaultdict(lambda: {
            'numero': '',
            'tipo': '',
            'fecha': '',
            'productos': [],
            'total': 0.0,
            'cantidad_total': 0,
            'cliente': {
                'tipo_documento': '',
                'numero_documento': '',
                'nombre': ''
            }
        })

        for fila in filas:
            comp_id = fila['factura_id']
            comprobante = comprobantes[comp_id]

            comprobante['numero'] = fila['numero_factura']
            comprobante['tipo'] = fila['tipo_documento']
            comprobante['fecha'] = fila['fecha_emision'].strftime('%d/%m/%Y') if isinstance(fila['fecha_emision'], datetime) else str(fila['fecha_emision'])

            comprobante['cliente']['tipo_documento'] = fila['cliente_tipo_documento']
            comprobante['cliente']['numero_documento'] = fila['cliente_numero_documento']
            comprobante['cliente']['nombre'] = fila['cliente_nombre']

            comprobante['productos'].append({
                'nombre': fila['producto_nombre'],
                'cantidad': int(fila['cantidad']),
                'precio_unitario': float(fila['precio_unitario']),
                'subtotal': float(fila['subtotal'])
            })

            comprobante['total'] += float(fila['subtotal'])
            comprobante['cantidad_total'] += int(fila['cantidad'])

        todos = list(comprobantes.values())

        if numero:
            todos = [c for c in todos if numero in c['numero'].lower()]

        if fecha:
            try:
                fecha_formateada = datetime.strptime(fecha, '%Y-%m-%d').strftime('%d/%m/%Y')
                todos = [c for c in todos if c['fecha'] == fecha_formateada]
            except ValueError:
                flash('Formato de fecha inválido. Usa YYYY-MM-DD.', 'warning')

        if tipo == 'factura':
            boletas = []
            facturas = [c for c in todos if c['tipo'].lower() == 'factura']
        elif tipo == 'boleta':
            facturas = []
            boletas = [c for c in todos if c['tipo'].lower() == 'boleta']
        else:
            facturas = [c for c in todos if c['tipo'].lower() == 'factura']
            boletas = [c for c in todos if c['tipo'].lower() == 'boleta']

        cursor.close()
        conn.close()

        return render_template('productos/vendidos.html', facturas=facturas, boletas=boletas, tipo=tipo)

    except Exception as e:
        flash(f'Error al obtener comprobantes vendidos: {e}', 'danger')
        return redirect(url_for('productos.mostrar_y_agregar_productos'))



    
@productos_bp.route('/vendidos/reporte-diario')
def reporte_diario_productos():
    try:
        fecha_inicio = request.args.get('fecha_inicio', '').strip()
        fecha_fin = request.args.get('fecha_fin', '').strip()

        if not fecha_inicio:
            # Si no hay fecha de inicio, mostrar modal
            return render_template('productos/reporte_diario.html', productos=None)

        # Convertir fechas
        fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d')

        if fecha_fin:
            fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d') + timedelta(days=1)
        else:
            # Si no se da fecha_fin, solo ese día
            fecha_fin_dt = fecha_inicio_dt + timedelta(days=1)

        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT 
                p.nombre AS producto_nombre,
                c.nombre AS categoria_nombre,
                SUM(df.cantidad) AS cantidad_total,
                SUM(df.cantidad * df.precio_unitario) AS monto_total
            FROM detalle_factura df
            JOIN productos p ON df.producto_id = p.id
            JOIN categoria c ON p.categoria_id = c.id
            JOIN facturas f ON df.factura_id = f.id
            WHERE f.fecha_emision >= %s AND f.fecha_emision < %s
            GROUP BY p.nombre, c.nombre
            ORDER BY p.nombre ASC
        """, (fecha_inicio_dt, fecha_fin_dt))

        rows = cursor.fetchall()
        cursor.close()
        conn.close()

        productos = []
        cantidad_total_general = 0
        total_general = 0.0

        for row in rows:
            cantidad = int(row['cantidad_total'])
            monto = float(row['monto_total'])
            precio_unitario = monto / cantidad if cantidad > 0 else 0

            productos.append({
                'nombre': row['producto_nombre'],
                'categoria': row['categoria_nombre'],
                'cantidad_vendida': cantidad,
                'precio_unitario': precio_unitario,
                'subtotal': monto
            })

            cantidad_total_general += cantidad
            total_general += monto

        return render_template(
            'productos/reporte_diario.html',
            productos=productos,
            fecha_inicio=fecha_inicio,
            fecha_fin=fecha_fin or fecha_inicio,
            cantidad_total_general=cantidad_total_general,
            total_general=total_general
        )

    except Exception as e:
        flash(f'Error al generar el reporte: {e}', 'danger')
        return redirect(url_for('productos.productos_vendidos'))


@productos_bp.route('/vendidos/reporte-diario/exportar')
def exportar_reporte_diario_productos_excel():
    fecha_inicio = request.args.get('fecha_inicio', '').strip()
    fecha_fin = request.args.get('fecha_fin', '').strip()

    if not fecha_inicio:
        flash('Debes especificar al menos una fecha de inicio.', 'warning')
        return redirect(url_for('productos.reporte_diario_productos'))

    fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d')

    if fecha_fin:
        fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d') + timedelta(days=1)
    else:
        fecha_fin_dt = fecha_inicio_dt + timedelta(days=1)

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT 
            p.nombre AS producto_nombre,
            c.nombre AS categoria_nombre,
            SUM(df.cantidad) AS cantidad_total,
            SUM(df.cantidad * df.precio_unitario) AS monto_total
        FROM detalle_factura df
        JOIN productos p ON df.producto_id = p.id
        JOIN categoria c ON p.categoria_id = c.id
        JOIN facturas f ON df.factura_id = f.id
        WHERE f.fecha_emision >= %s AND f.fecha_emision < %s
        GROUP BY p.nombre, c.nombre
        ORDER BY p.nombre ASC
    """, (fecha_inicio_dt, fecha_fin_dt))

    rows = cursor.fetchall()
    cursor.close()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = f"Reporte {fecha_inicio} — {fecha_fin or fecha_inicio}"

    ws.append(["Producto", "Categoría", "Cantidad Vendida", "Precio Unitario (S/)", "Subtotal (S/)"])

    for row in rows:
        cantidad = row['cantidad_total']
        monto = row['monto_total']
        precio_unitario = monto / cantidad if cantidad > 0 else 0
        ws.append([
            row['producto_nombre'],
            row['categoria_nombre'],
            cantidad,
            round(precio_unitario, 2),
            round(monto, 2)
        ])

    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = make_response(output.read())
    filename = f"reporte_{fecha_inicio}_a_{fecha_fin or fecha_inicio}.xlsx"
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    response.headers["Content-type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return response


@productos_bp.route('/productos/reportes-mensuales')
def productos_reportes_mensuales():
    mes_actual = datetime.now().strftime('%Y-%m')  # formato '2025-06'

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT 
            f.tipo_documento,
            f.numero_factura,
            f.fecha_emision,
            p.nombre AS producto_nombre,
            df.cantidad,
            df.precio_unitario,
            (df.cantidad * df.precio_unitario) AS subtotal
        FROM detalle_factura df
        JOIN productos p ON df.producto_id = p.id
        JOIN facturas f ON df.factura_id = f.id
        ORDER BY f.fecha_emision DESC
    """)

    filas = cursor.fetchall()

    reportes_mensuales = defaultdict(lambda: {
        'comprobantes': [],
        'total_mensual': 0.0,
        'productos_total': 0
    })

    comprobantes_por_mes = defaultdict(dict)
    productos_por_mes = defaultdict(int)
    ingresos_por_mes = defaultdict(float)
    tipo_comprobantes = {'factura': 0, 'boleta': 0}

    for fila in filas:
        fecha = fila['fecha_emision']
        mes = fecha.strftime('%B %Y').capitalize()
        clave_comprobante = f"{fila['tipo_documento']}-{fila['numero_factura']}-{fecha.strftime('%d/%m/%Y')}"

        if clave_comprobante not in comprobantes_por_mes[mes]:
            comprobantes_por_mes[mes][clave_comprobante] = {
                'tipo': fila['tipo_documento'],
                'numero': fila['numero_factura'],
                'fecha': fecha.strftime('%d/%m/%Y'),
                'productos': [],
                'cantidad_total': 0,
                'total': 0.0
            }

        comprobante = comprobantes_por_mes[mes][clave_comprobante]
        comprobante['productos'].append({
            'nombre': fila['producto_nombre'],
            'cantidad': int(fila['cantidad']),
            'subtotal': float(fila['subtotal'])
        })
        comprobante['cantidad_total'] += int(fila['cantidad'])
        comprobante['total'] += float(fila['subtotal'])

        # Actualizar resumen mensual
        reportes_mensuales[mes]['total_mensual'] += float(fila['subtotal'])
        reportes_mensuales[mes]['productos_total'] += int(fila['cantidad'])

        productos_por_mes[mes] += int(fila['cantidad'])
        ingresos_por_mes[mes] += float(fila['subtotal'])

        tipo = fila['tipo_documento'].lower()
        if tipo in tipo_comprobantes:
            tipo_comprobantes[tipo] += 1

    # Agregar comprobantes organizados a reportes
    for mes, comprobantes in comprobantes_por_mes.items():
        reportes_mensuales[mes]['comprobantes'] = list(comprobantes.values())

    cursor.execute("""
    SELECT p.nombre, SUM(df.cantidad) as total_vendido
    FROM detalle_factura df
    JOIN productos p ON df.producto_id = p.id
    JOIN facturas f ON df.factura_id = f.id
    WHERE DATE_FORMAT(f.fecha_emision, '%%Y-%%m') = %s
    GROUP BY p.nombre
    ORDER BY total_vendido DESC
    LIMIT 5
    """, (mes_actual,))


    top_productos = cursor.fetchall()
    cursor.close()
    conn.close()

    top_nombres = [row['nombre'] for row in top_productos]
    top_cantidades = [int(row['total_vendido']) for row in top_productos]

    meses_ordenados = sorted(reportes_mensuales.keys())
    labels = meses_ordenados
    productos_data = [productos_por_mes[mes] for mes in labels]
    ingresos_data = [ingresos_por_mes[mes] for mes in labels]

    return render_template('productos/reporte_mensual.html',
                           reportes=reportes_mensuales,
                           labels=labels,
                           productos_por_mes=productos_data,
                           ingresos_por_mes=ingresos_data,
                           comprobantes=tipo_comprobantes,
                           top_nombres=top_nombres,
                           top_cantidades=top_cantidades)

@productos_bp.route('/reporte_mensual/exportar')
def exportar_reporte_mensual_excel():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT 
                DATE_FORMAT(f.fecha_emision, '%%Y-%%m') AS mes,
                f.tipo_documento,
                SUM(df.cantidad * df.precio_unitario) AS total_ventas,
                SUM(df.cantidad) AS total_productos,
                COUNT(DISTINCT f.id) AS total_comprobantes
            FROM detalle_factura df
            JOIN productos p ON df.producto_id = p.id
            JOIN facturas f ON df.factura_id = f.id
            GROUP BY mes, f.tipo_documento
            ORDER BY mes DESC
        """)

        resumen = cursor.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Mensual"

        ws.append(["Mes", "Tipo", "Total Comprobantes", "Total Productos", "Total Ventas (S/)"])
        for r in resumen:
            ws.append([
                r['mes'],
                r['tipo_documento'].capitalize(),
                r['total_comprobantes'],
                r['total_productos'],
                round(r['total_ventas'], 2)
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(output, download_name="reporte_mensual.xlsx", as_attachment=True)

    except Exception as e:
        flash(f'Error al exportar reporte mensual: {e}', 'danger')
        return redirect(url_for('productos.reporte_mensual'))

@productos_bp.route('/exportar_comprobante/<string:numero>')
def exportar_comprobante_excel(numero):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT 
                f.numero_factura,
                f.fecha_emision,
                f.tipo_documento,
                p.nombre AS producto_nombre,
                df.cantidad,
                df.precio_unitario,
                (df.cantidad * df.precio_unitario) AS subtotal
            FROM detalle_factura df
            JOIN productos p ON df.producto_id = p.id
            JOIN facturas f ON df.factura_id = f.id
            WHERE f.numero_factura = %s
        """, (numero,))

        filas = cursor.fetchall()
        if not filas:
            flash('Comprobante no encontrado.', 'danger')
            return redirect(url_for('productos.productos_vendidos'))

        comprobante = filas[0]
        tipo = comprobante['tipo_documento']
        data = []
        for fila in filas:
            data.append({
                'Producto': fila['producto_nombre'],
                'Cantidad': fila['cantidad'],
                'Precio Unitario': f"S/ {fila['precio_unitario']:.2f}",
                'Subtotal': f"S/ {fila['subtotal']:.2f}"
            })

        df = pd.DataFrame(data)
        df.insert(0, f'{tipo} N°', comprobante['numero_factura'])
        df.insert(1, 'Fecha', comprobante['fecha_emision'].strftime('%d/%m/%Y'))

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=tipo, index=False)

        output.seek(0)
        filename = f"{tipo}_{numero}.xlsx"

        cursor.close()
        conn.close()

        return send_file(
            output,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        flash(f'Error al exportar comprobante: {e}', 'danger')
        return redirect(url_for('productos.productos_vendidos'))
    

@productos_bp.route('/exportar_todos_comprobantes')
def exportar_todos_comprobantes_excel():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        cursor.execute("""
            SELECT 
                f.numero_factura AS numero,
                f.fecha_emision AS fecha,
                f.tipo_documento,
                p.nombre AS producto,
                df.cantidad,
                df.precio_unitario,
                (df.cantidad * df.precio_unitario) AS subtotal
            FROM detalle_factura df
            JOIN productos p ON df.producto_id = p.id
            JOIN facturas f ON df.factura_id = f.id
            ORDER BY f.fecha_emision ASC
        """)
        rows = cursor.fetchall()
        cursor.close()
        conn.close()

        if not rows:
            flash('No hay comprobantes para exportar.', 'warning')
            return redirect(url_for('productos.productos_vendidos'))

        data = []
        for r in rows:
            data.append({
                "Tipo": r["tipo_documento"].capitalize(),  # Factura o Boleta
                "Número": r["numero"],
                "Fecha": r["fecha"].strftime('%d/%m/%Y'),
                "Producto": r["producto"],
                "Cantidad": r["cantidad"],
                "Precio Unitario (S/)": round(r["precio_unitario"], 2),
                "Subtotal (S/)": round(r["subtotal"], 2)
            })

        df = pd.DataFrame(data)

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Comprobantes', index=False)
            wb = writer.book
            ws = writer.sheets['Comprobantes']

            # Estilos
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            border = Border(
                left=Side(border_style="thin", color="000000"),
                right=Side(border_style="thin", color="000000"),
                top=Side(border_style="thin", color="000000"),
                bottom=Side(border_style="thin", color="000000")
            )
            align_center = Alignment(horizontal="center", vertical="center")

            for col in range(1, len(df.columns) + 1):
                col_letter = get_column_letter(col)
                ws.column_dimensions[col_letter].width = 20
                cell = ws[f"{col_letter}1"]
                cell.fill = header_fill
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = align_center
                cell.border = border

            for row in range(2, ws.max_row + 1):
                for col in range(1, len(df.columns) + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.alignment = align_center
                    cell.border = border
                    if row % 2 == 0:
                        cell.fill = zebra_fill

        output.seek(0)
        return send_file(
            output,
            as_attachment=True,
            download_name="Comprobantes_Vendidos.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        flash(f'Error al exportar comprobantes: {e}', 'danger')
        return redirect(url_for('productos.productos_vendidos'))