from flask import Blueprint, json, render_template, request, redirect, url_for, flash, send_file,jsonify
from data import get_db_connection
from datetime import datetime
from decimal import Decimal
from io import BytesIO
import os
import requests
from fpdf import FPDF
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import mysql.connector  
import pymysql

#token para consumir api sunat
TOKEN = "eyJ0eXAiOiJKV1QiLCJhbGciOiJIUzI1NiJ9.eyJlbWFpbCI6ImFsZXhzbGNkcHI1N0BnbWFpbC5jb20ifQ.g59rpjSgG3HC8bhWBaVdHCRcE7zXsBMQAYKapybT_84"


facturas_bp = Blueprint('facturas', __name__, url_prefix='/facturas')

@facturas_bp.route('/')
def listar_facturas():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT f.id, f.numero_factura, f.tipo_documento, c.nombre AS cliente, f.fecha_emision, f.total
        FROM facturas f
        JOIN clientes c ON f.cliente_documento_numero = c.documento_numero
        ORDER BY f.fecha_emision DESC
    """)
    facturas = cursor.fetchall()
    cursor.close()
    conn.close()
    return render_template('facturas/listado_facturas.html', facturas=facturas)

# Ruta para la API de búsqueda de clientes
@facturas_bp.route('/api/buscar-cliente', methods=['GET'])
def buscar_cliente_api():
    documento = request.args.get('documento')
    
    if not documento:
        return jsonify({"error": "Documento requerido"}), 400

    conn = None
    cursor = None
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        print(f"Buscando cliente con documento: {documento}")
        
        cursor.execute("""
            SELECT documento_numero, nombre, direccion, telefono, estado 
            FROM clientes 
            WHERE documento_numero = %s
            LIMIT 1
        """, (documento,))
        
        # pymysql con DictCursor ya devuelve diccionarios
        cliente = cursor.fetchone()
        print(f"Resultado de la consulta: {cliente}")
        
        if cliente:
            # Convertir el estado a booleano
            cliente['estado'] = bool(cliente['estado'])
            print(f"Cliente encontrado: {cliente}")
            return jsonify(cliente)
        else:
            return jsonify({
                "encontrado": False,
                "mensaje": "Cliente no registrado"
            }), 200
            
    except pymysql.Error as e:
        print(f"[ERROR DB] {str(e)}")
        return jsonify({
            "error": "Error de base de datos",
            "detalles": str(e)
        }), 500
    except Exception as e:
        print(f"[ERROR] {type(e).__name__}: {str(e)}")
        return jsonify({
            "error": "Error inesperado",
            "detalles": str(e) if str(e) else "Error desconocido"
        }), 500
    finally:
        try:
            if cursor:
                cursor.close()
        except Exception as e:
            print(f"Error cerrando cursor: {str(e)}")
        
        try:
            if conn:
                conn.close()
        except Exception as e:
            print(f"Error cerrando conexión: {str(e)}")

@facturas_bp.route('/api/detalle_factura')
def api_detalle_factura():
    factura_id = request.args.get('factura_id')

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT 
            df.id AS detalle_id,
            df.producto_id AS id,
            p.nombre,
            CAST(df.cantidad AS DECIMAL(10,2)) AS cantidad,
            CAST(df.precio_unitario AS DECIMAL(10,2)) AS precio_unitario,
            CAST(p.precio AS DECIMAL(10,2)) AS nuevo_precio
        FROM detalle_factura df
        JOIN productos p ON df.producto_id = p.id
        WHERE df.factura_id = %s
    """, (factura_id, ))

    productos = cursor.fetchall()

    # Convertimos los decimales a float para evitar problemas de serialización
    productos_sanitizados = []
    for p in productos:
        productos_sanitizados.append({
            "detalle_id": p["detalle_id"],
            "id": p["id"],
            "nombre": p["nombre"],
            "cantidad": float(p["cantidad"]),
            "precio_unitario": float(p["precio_unitario"]),
            "nuevo_precio": float(p["nuevo_precio"])
        })

    cursor.close()
    conn.close()

    return jsonify({"productos": productos_sanitizados})


@facturas_bp.route('/api/facturas_por_cliente')
def api_facturas_por_cliente():
    documento = request.args.get('documento')

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT f.id, f.numero_factura, f.total, c.nombre AS cliente
        FROM facturas f
        JOIN clientes c ON f.cliente_documento_numero = c.documento_numero
        WHERE c.documento_numero = %s
    """, (documento, ))

    facturas = cursor.fetchall()

    cursor.close()
    conn.close()

    return jsonify({
        "success": bool(facturas),
        "facturas": facturas
    })


@facturas_bp.route('/api/buscar_factura')
def api_buscar_factura():
    criterio = request.args.get('criterio')

    conn = get_db_connection()
    cursor = conn.cursor()

    # Buscar por número exacto, solo FACTURA
    cursor.execute("""
        SELECT f.id, f.numero_factura, f.total, c.nombre AS cliente
        FROM facturas f
        JOIN clientes c ON f.cliente_documento_numero = c.documento_numero
        WHERE f.numero_factura = %s AND f.tipo_documento = 'FACTURA'
    """, (criterio, ))

    facturas = cursor.fetchall()

    # Si no encontró por número, buscar por RUC/DNI, solo FACTURA
    if not facturas:
        cursor.execute("""
            SELECT f.id, f.numero_factura, f.total, c.nombre AS cliente
            FROM facturas f
            JOIN clientes c ON f.cliente_documento_numero = c.documento_numero
            WHERE c.documento_numero = %s AND f.tipo_documento = 'FACTURA'
        """, (criterio, ))

        facturas = cursor.fetchall()

    cursor.close()
    conn.close()

    return jsonify({
        "success": bool(facturas),
        "facturas": facturas
    })



@facturas_bp.route('/api/precio_actual_producto')
def api_precio_actual_producto():
    producto_id = request.args.get('producto_id')
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Usamos la columna correcta: precio
    cursor.execute("SELECT precio FROM productos WHERE id = %s", (producto_id,))
    prod = cursor.fetchone()
    
    cursor.close()
    conn.close()

    if prod:
        return jsonify({"success": True, "precio": float(prod['precio'])})
    else:
        return jsonify({"success": False, "error": "Producto no encontrado"})

@facturas_bp.route('/nueva', methods=['GET', 'POST'])
def nueva_factura():
    conn = get_db_connection()
    cursor = conn.cursor()

    if request.method == 'POST':
        cliente_documento = request.form.get('cliente_documento')
        tipo_documento = request.form.get('tipo_documento')
        numero_documento = request.form.get('numero_documento')
        productos = request.form.getlist('producto_id')
        cantidades = request.form.getlist('cantidad')

        if not cliente_documento or not productos or not cantidades or not tipo_documento or not numero_documento:
            flash("Complete todos los campos obligatorios", "danger")
            return redirect(url_for('facturas.nueva_factura'))

        try:
            cantidades = [int(c) for c in cantidades]
            productos = [int(p) for p in productos]
            if any(c <= 0 for c in cantidades):
                flash("Las cantidades deben ser mayores a cero", "warning")
                return redirect(url_for('facturas.nueva_factura'))
        except ValueError:
            flash("Datos inválidos en productos/cantidades", "danger")
            return redirect(url_for('facturas.nueva_factura'))

        usuario_id = 1
        usuario_username = 'admin'

        total = 0
        detalles = []

        try:
            for producto_id, cantidad in zip(productos, cantidades):
                cursor.execute("SELECT id, nombre, precio, stock FROM productos WHERE id = %s", (producto_id,))
                producto = cursor.fetchone()

                if not producto:
                    flash(f"Producto ID {producto_id} no existe", "danger")
                    return redirect(url_for('facturas.nueva_factura'))

                if producto['stock'] < cantidad:
                    flash(f"Stock insuficiente para {producto['nombre']}", "warning")
                    return redirect(url_for('facturas.nueva_factura'))

                precio_unitario_con_igv = float(producto['precio'])
                subtotal_con_igv = precio_unitario_con_igv * cantidad
                total += subtotal_con_igv

                detalles.append((
                    producto_id,
                    cantidad,
                    precio_unitario_con_igv,
                    subtotal_con_igv
                ))

            # ===== DEBUG: imprimir lo que se va a guardar =====
            print("======= DATOS A GUARDAR EN FACTURA =======")
            print(f"Cliente: {cliente_documento}")
            print(f"Tipo de Documento: {tipo_documento}")
            print(f"Número de Documento: {numero_documento}")
            print(f"Usuario: {usuario_username}")
            print(f"TOTAL (con IGV): S/ {total:.2f}")
            print("----------- DETALLES -----------")
            for idx, (pid, cantidad, precio_unitario, subtotal) in enumerate(detalles, start=1):
                print(f"{idx}. Producto ID: {pid} | Cantidad: {cantidad} | Precio Unitario: S/ {precio_unitario:.2f} | Subtotal: S/ {subtotal:.2f}")
            print("==========================================")

            cursor.execute("""
                INSERT INTO facturas (
                    numero_factura, 
                    tipo_documento,
                    cliente_documento_numero, 
                    usuario_id, 
                    usuario_username, 
                    fecha_emision, 
                    total
                ) VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (
                numero_documento,
                tipo_documento,
                cliente_documento,
                usuario_id,
                usuario_username,
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                total
            ))

            factura_id = cursor.lastrowid

            for detalle in detalles:
                cursor.execute("""
                    INSERT INTO detalle_factura (
                        factura_id, 
                        producto_id, 
                        cantidad, 
                        precio_unitario, 
                        subtotal
                    ) VALUES (%s, %s, %s, %s, %s)
                """, (factura_id, *detalle))

                cursor.execute("UPDATE productos SET stock = stock - %s WHERE id = %s", (detalle[1], detalle[0]))

            conn.commit()
            flash("Documento registrado exitosamente!", "success")
            return redirect(url_for('facturas.listar_facturas'))

        except mysql.connector.Error as e:
            conn.rollback()
            flash(f"Error de base de datos: {str(e)}", "danger")
        except Exception as e:
            conn.rollback()
            flash(f"Error al procesar documento: {str(e)}", "danger")
        finally:
            cursor.close()
            conn.close()

        return redirect(url_for('facturas.nueva_factura'))

    # GET
    try:
        cursor.execute("SELECT id, nombre, precio, stock, codigo_producto FROM productos WHERE stock > 0")
        productos = cursor.fetchall()
        return render_template('facturas/nueva_factura.html', productos=productos)
    finally:
        cursor.close()
        conn.close()

@facturas_bp.route('/generar_pdf_factura/<string:numero>')
def generar_pdf_factura(numero):
    return _generar_pdf_comprobante(numero, 'factura')

@facturas_bp.route('/generar_pdf_boleta/<string:numero>')
def generar_pdf_boleta(numero):
    return _generar_pdf_comprobante(numero, 'boleta')

def _generar_pdf_comprobante(numero, tipo_doc):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT f.numero_factura, f.fecha_emision, f.tipo_documento,
                   c.nombre AS cliente, c.documento_tipo, c.documento_numero, c.direccion, c.telefono,
                   p.codigo_producto, p.nombre AS producto,
                   df.cantidad, df.precio_unitario, (df.cantidad * df.precio_unitario) AS subtotal
            FROM facturas f
            JOIN detalle_factura df ON f.id = df.factura_id
            JOIN productos p ON df.producto_id = p.id
            JOIN clientes c ON f.cliente_documento_numero = c.documento_numero
            WHERE f.numero_factura = %s AND f.tipo_documento = %s
        """, (numero, tipo_doc))
        filas = cursor.fetchall()
        cursor.close()
        conn.close()

        if not filas:
            flash(f"{tipo_doc.capitalize()} no encontrado", "warning")
            return redirect(url_for('productos.productos_vendidos'))

        empresa = {
            'nombre': 'GOLOCINA SAC',
            'ruc': '10404129207',
            'direccion': 'CAR.CENTRAL KM. 6.5 INT. 258 OTR. CENTRO COMERCIAL LAS BRISAS DE ATE LIMA - LIMA - ATE',
            'telefono': '(01) 123-4567',
            'email': 'contacto@gmail.com',
            'logo': os.path.join(os.path.dirname(__file__), '..', 'static', 'img', 'logo.png')
        }

        f0 = filas[0]
        cliente = {
            'nombre': f0['cliente'],
            'documento': f"{f0['documento_tipo']} {f0['documento_numero']}",
            'direccion': f0['direccion'] or '-',
            'telefono': f0['telefono'] or '-'
        }

        productos = []
        total = Decimal('0.00')
        for fila in filas:
            productos.append({
                'codigo': fila['codigo_producto'],
                'nombre': fila['producto'],
                'cantidad': str(fila['cantidad']),
                'precio': f"S/ {Decimal(fila['precio_unitario']):.2f}",
                'subtotal': f"S/ {Decimal(fila['subtotal']):.2f}"
            })
            total += Decimal(fila['subtotal'])

        subtotal_sin_igv = (total / Decimal('1.18')).quantize(Decimal('0.01'))
        igv = (total - subtotal_sin_igv).quantize(Decimal('0.01'))

        pdf = FPDF('P', 'mm', 'A4')
        pdf.add_page()
        pdf.set_margins(15, 15, 15)
        pdf.set_auto_page_break(True, 15)

        if os.path.exists(empresa['logo']):
            pdf.image(empresa['logo'], 15, 10, 30)
            pdf.set_y(40)
        else:
            pdf.set_y(20)

        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 8, empresa['nombre'], ln=1)
        pdf.set_font('Arial', '', 10)
        pdf.cell(0, 5, f"RUC: {empresa['ruc']}", ln=1)
        pdf.multi_cell(0, 5, empresa['direccion'])
        pdf.cell(0, 5, f"Tel: {empresa['telefono']} | Email: {empresa['email']}", ln=1)
        pdf.line(15, pdf.get_y() + 5, 195, pdf.get_y() + 5)
        pdf.ln(10)

        pdf.set_font('Arial', 'B', 16)
        pdf.set_text_color(0, 51, 102)
        pdf.cell(0, 10, f"{tipo_doc.upper()} ELECTRÓNICA", ln=1, align='C')
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, f"N° {numero}", ln=1, align='C')
        pdf.set_font('Arial', '', 10)
        pdf.set_text_color(0, 0, 0)
        pdf.cell(0, 5, f"Fecha: {f0['fecha_emision'].strftime('%d/%m/%Y')}", ln=1)
        pdf.ln(8)

        pdf.set_font('Arial', 'B', 12)
        pdf.cell(0, 8, "DATOS DEL CLIENTE", ln=1)
        pdf.set_fill_color(240, 240, 240)
        pdf.rect(15, pdf.get_y(), 180, 20, 'F')
        pdf.set_font('Arial', '', 10)
        pdf.set_xy(20, pdf.get_y() + 2)
        pdf.cell(0, 5, f"Nombre: {cliente['nombre']}", ln=1)
        pdf.set_x(20)
        pdf.cell(0, 5, f"Documento: {cliente['documento']}", ln=1)
        pdf.set_x(20)
        pdf.cell(0, 5, f"Dirección: {cliente['direccion']}", ln=1)
        pdf.set_x(20)
        pdf.cell(0, 5, f"Teléfono: {cliente['telefono']}", ln=1)
        pdf.ln(8)

        pdf.set_font('Arial', 'B', 12)
        pdf.cell(0, 8, "DETALLE DE PRODUCTOS", ln=1)
        pdf.set_fill_color(0, 51, 102)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font('Arial', 'B', 10)

        widths = [30, 70, 20, 30, 30]
        headers = ["Código", "Descripción", "Cant.", "P. Unitario", "Subtotal"]
        for i, header in enumerate(headers):
            pdf.cell(widths[i], 8, header, 1, 0, 'C', 1)
        pdf.ln()

        pdf.set_font('Arial', '', 10)
        pdf.set_text_color(0, 0, 0)
        fill = False
        for p in productos:
            pdf.set_fill_color(240, 240, 240) if fill else pdf.set_fill_color(255, 255, 255)
            pdf.cell(widths[0], 8, p['codigo'], 1, 0, 'L', 1)
            pdf.cell(widths[1], 8, p['nombre'], 1, 0, 'L', 1)
            pdf.cell(widths[2], 8, p['cantidad'], 1, 0, 'C', 1)
            pdf.cell(widths[3], 8, p['precio'], 1, 0, 'R', 1)
            pdf.cell(widths[4], 8, p['subtotal'], 1, 1, 'R', 1)
            fill = not fill

        pdf.set_font('Arial', 'B', 10)
        pdf.set_fill_color(220, 220, 220)
        pdf.cell(sum(widths[:4]), 8, "Subtotal (sin IGV):", 1, 0, 'R', 1)
        pdf.cell(widths[4], 8, f"S/ {subtotal_sin_igv:.2f}", 1, 1, 'R', 1)
        pdf.cell(sum(widths[:4]), 8, "IGV (18%):", 1, 0, 'R', 1)
        pdf.cell(widths[4], 8, f"S/ {igv:.2f}", 1, 1, 'R', 1)
        pdf.set_fill_color(0, 51, 102)
        pdf.set_text_color(255, 255, 255)
        pdf.cell(sum(widths[:4]), 8, "TOTAL:", 1, 0, 'R', 1)
        pdf.cell(widths[4], 8, f"S/ {total:.2f}", 1, 1, 'R', 1)

        pdf.ln(10)
        pdf.set_font('Arial', 'I', 8)
        pdf.set_text_color(100, 100, 100)
        pdf.cell(0, 5, "¡Gracias por su compra!", 0, 1, 'C')
        pdf.cell(0, 5, "Este documento es una representación impresa de un comprobante electrónico", 0, 1, 'C')

        pdf_output = BytesIO(pdf.output(dest='S').encode('latin1'))
        pdf_output.seek(0)
        return send_file(pdf_output, mimetype='application/pdf', as_attachment=True, download_name=f"{tipo_doc.capitalize()}_{numero}.pdf")

    except Exception as e:
        flash(f"Error al generar PDF: {e}", "danger")
        return redirect(url_for('productos.productos_vendidos'))


@facturas_bp.route('/borrar/<int:id>', methods=['POST'])
def borrar_facturas(id):
    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM facturas WHERE id = %s", (id,))
    factura = cursor.fetchone()

    if not factura:
        cursor.close()
        conn.close()
        flash("Factura no encontrada.", "danger")
        return redirect(url_for('facturas.listar_facturas'))

    try:
        cursor.execute("DELETE FROM detalle_factura WHERE factura_id = %s", (id,))
        cursor.execute("DELETE FROM facturas WHERE id = %s", (id,))
        conn.commit()
        flash("Factura eliminada correctamente.", "success")

    except mysql.connector.Error as e:
        conn.rollback()
        flash(f"Error al eliminar la factura: {e}", "danger")

    finally:
        cursor.close()
        conn.close()

    return redirect(url_for('facturas.listar_facturas'))

@facturas_bp.route('/nueva-nota', methods=['GET'])
def nueva_nota():
    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT f.id, f.numero_factura, c.nombre AS cliente, f.total 
        FROM facturas f
        JOIN clientes c ON f.cliente_documento_numero = c.documento_numero
        WHERE f.tipo_documento = 'factura'
        ORDER BY f.fecha_emision DESC
    """)
    facturas = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template('facturas/Nota_Credito_debito.html', facturas=facturas)


# --- 2) Registrar nota ---
@facturas_bp.route('/registrar-nota', methods=['POST'])
def registrar_nota():
    factura_id = request.form.get('factura_id')
    tipo_nota = request.form.get('tipo_nota')  # 'credito' o 'debito'
    motivo = request.form.get('motivo')

    if not factura_id or not tipo_nota or not motivo:
        return jsonify({"success": False, "error": "Datos incompletos en el formulario."})

    conn = get_db_connection()
    cursor = conn.cursor()

    # Obtener datos de factura y cliente
    cursor.execute("""
        SELECT f.id, f.total, f.cliente_documento_numero, c.nombre 
        FROM facturas f 
        JOIN clientes c ON f.cliente_documento_numero = c.documento_numero
        WHERE f.id = %s
    """, (factura_id,))
    factura = cursor.fetchone()

    if not factura:
        cursor.close()
        conn.close()
        return jsonify({"success": False, "error": "Factura no encontrada"})

    # Generar código correlativo
    cursor.execute("SELECT COUNT(*) AS total FROM notas_credito_debito")
    count = cursor.fetchone()['total'] + 1
    codigo = f"NCD-{count:06d}"

    monto = Decimal('0.00')
    detalle_productos = []

    if tipo_nota == 'credito':
        detalle_json = request.form.get('detalle')
        if not detalle_json:
            cursor.close()
            conn.close()
            return jsonify({"success": False, "error": "No se enviaron productos para la nota de crédito."})

        try:
            detalle = json.loads(detalle_json)
        except Exception as e:
            cursor.close()
            conn.close()
            return jsonify({"success": False, "error": "Detalle de productos inválido."})

        if not detalle:
            cursor.close()
            conn.close()
            return jsonify({"success": False, "error": "La nota de crédito no contiene productos."})

        for item in detalle:
            producto_id = item.get('producto_id')
            cantidad = Decimal(item.get('cantidad', 0))
            precio_unitario = Decimal(item.get('precio_unitario', 0))
            subtotal = cantidad * precio_unitario

            detalle_productos.append({
                'producto_id': producto_id,
                'cantidad': float(cantidad),
                'precio_unitario': float(precio_unitario),
                'subtotal': float(subtotal)
            })
            monto += subtotal

        cursor.execute("UPDATE facturas SET estado = 'anulada' WHERE id = %s", (factura_id,))

    elif tipo_nota == 'debito':
        cursor.execute("""
            SELECT df.producto_id, df.cantidad, 
                   CAST(p.precio AS DECIMAL(10,2)) AS nuevo_precio, 
                   CAST(df.precio_unitario AS DECIMAL(10,2)) AS antiguo_precio
            FROM detalle_factura df
            JOIN productos p ON df.producto_id = p.id
            WHERE df.factura_id = %s
        """, (factura_id,))
        productos = cursor.fetchall()

        if not productos:
            cursor.close()
            conn.close()
            return jsonify({"success": False, "error": "No se encontraron productos asociados a la factura."})

        for prod in productos:
            nuevo_precio = Decimal(prod['nuevo_precio'])
            antiguo_precio = Decimal(prod['antiguo_precio'])
            cantidad = Decimal(prod['cantidad'])

            if nuevo_precio != antiguo_precio:
                diferencia = (nuevo_precio - antiguo_precio) * cantidad
                detalle_productos.append({
                    'producto_id': prod['producto_id'],
                    'cantidad': float(cantidad),
                    'precio_unitario': float(nuevo_precio),
                    'subtotal': float(diferencia)
                })
                monto += diferencia

        if not detalle_productos:
            cursor.close()
            conn.close()
            return jsonify({
                "success": False,
                "error": "No hay diferencias de precio que generen una nota de débito. Verifica los productos."
            })

    # Insertar nota
    cursor.execute("""
        INSERT INTO notas_credito_debito (factura_id, tipo_nota, motivo, monto, codigo, fecha_emision, estado)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
    """, (
        factura_id, tipo_nota, motivo, float(monto), codigo, datetime.now(), 'activo'
    ))
    nota_id = cursor.lastrowid

    # Insertar detalle de nota
    for item in detalle_productos:
        cursor.execute("""
            INSERT INTO detalle_nota (nota_id, producto_id, cantidad, precio_unitario, subtotal)
            VALUES (%s, %s, %s, %s, %s)
        """, (
            nota_id, item['producto_id'], item['cantidad'], item['precio_unitario'], item['subtotal']
        ))

    conn.commit()
    cursor.close()
    conn.close()

    return jsonify({"success": True, "nota_id": nota_id})

# --- 3) Generar PDF ---
@facturas_bp.route('/generar-pdf-nota/<int:nota_id>')
def generar_pdf_nota(nota_id):

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT n.*, f.numero_factura, c.nombre AS cliente_nombre, c.documento_numero, c.documento_tipo
        FROM notas_credito_debito n
        JOIN facturas f ON n.factura_id = f.id
        JOIN clientes c ON f.cliente_documento_numero = c.documento_numero
        WHERE n.id = %s
    """, (nota_id,))
    nota = cursor.fetchone()

    if not nota:
        cursor.close()
        conn.close()
        return "Nota no encontrada", 404

    productos = []
    if nota['tipo_nota'] == 'credito':
        # Obtener productos de la nota (detalle_nota) para crédito también
        cursor.execute("""
            SELECT dn.*, p.nombre
            FROM detalle_nota dn
            JOIN productos p ON dn.producto_id = p.id
            WHERE dn.nota_id = %s
        """, (nota_id,))
        productos = cursor.fetchall()
    else:
        cursor.execute("""
            SELECT dn.*, p.nombre
            FROM detalle_nota dn
            JOIN productos p ON dn.producto_id = p.id
            WHERE dn.nota_id = %s
        """, (nota_id,))
        productos = cursor.fetchall()

    cursor.close()
    conn.close()

    subtotal_con_igv = round(sum(float(p['subtotal']) for p in productos), 2)
    subtotal_sin_igv = round(subtotal_con_igv / 1.18, 2)
    igv = round(subtotal_con_igv - subtotal_sin_igv, 2)
    total = subtotal_con_igv

    empresa = {
        'nombre': 'GOLOCINA SAC',
        'ruc': '10404129207',
        'direccion': 'CAR.CENTRAL KM. 6.5 INT. 258 OTR. CENTRO COMERCIAL LAS BRISAS DE ATE LIMA - LIMA - ATE',
        'telefono': '(01) 123-4567',
        'email': 'contacto@gmail.com',
        'logo': os.path.join(os.path.dirname(__file__), '..', 'static', 'img', 'logo.png')
    }

    pdf = FPDF(orientation='P', unit='mm', format='A4')
    pdf.add_page()
    pdf.set_margins(15, 15, 15)
    pdf.set_auto_page_break(auto=True, margin=15)

    if os.path.exists(empresa['logo']):
        pdf.image(empresa['logo'], x=15, y=10, w=30)
        pdf.set_y(40)
    else:
        pdf.set_y(20)

    pdf.set_font('Arial', 'B', 14)
    pdf.cell(0, 8, empresa['nombre'], ln=1)
    pdf.set_font('Arial', '', 10)
    pdf.cell(0, 5, f"RUC: {empresa['ruc']}", ln=1)
    pdf.multi_cell(0, 5, empresa['direccion'])
    pdf.cell(0, 5, f"Tel: {empresa['telefono']} | Email: {empresa['email']}", ln=1)
    pdf.ln(5)

    pdf.set_font('Arial', 'B', 16)
    pdf.set_text_color(0, 51, 102)
    tipo_nota_str = "NOTA DE CRÉDITO" if nota['tipo_nota'] == 'credito' else "NOTA DE DÉBITO"
    pdf.cell(0, 10, f"{tipo_nota_str} ELECTRÓNICA", ln=1, align='C')
    pdf.set_font('Arial', 'B', 14)
    pdf.cell(0, 10, f"N° {nota['codigo']}", ln=1, align='C')

    pdf.set_font('Arial', '', 10)
    pdf.set_text_color(0, 0, 0)
    pdf.cell(0, 5, f"Fecha de Emisión: {nota['fecha_emision'].strftime('%Y-%m-%d %H:%M:%S')}", ln=1)
    pdf.cell(0, 5, f"Factura: {nota['numero_factura']}", ln=1)
    pdf.cell(0, 5, f"Cliente: {nota['cliente_nombre']} - {nota['documento_tipo']} {nota['documento_numero']}", ln=1)
    pdf.cell(0, 5, f"Motivo: {nota['motivo']}", ln=1)
    pdf.ln(5)

    pdf.set_font('Arial', 'B', 12)
    pdf.cell(0, 8, "DETALLE DE PRODUCTOS", ln=1)

    pdf.set_fill_color(0, 51, 102)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font('Arial', 'B', 10)

    widths = [100, 20, 30, 30]
    headers = ["Descripción", "Cantidad", "P. Unitario", "Subtotal"]
    for i, header in enumerate(headers):
        pdf.cell(widths[i], 8, header, border=1, align='C', fill=True)
    pdf.ln()

    pdf.set_font('Arial', '', 10)
    pdf.set_text_color(0, 0, 0)
    fill = False

    if productos:
        for p in productos:
            pdf.set_fill_color(240, 240, 240) if fill else pdf.set_fill_color(255, 255, 255)
            pdf.cell(widths[0], 8, p['nombre'], border=1, fill=fill)
            pdf.cell(widths[1], 8, str(p['cantidad']), border=1, align='C', fill=fill)
            pdf.cell(widths[2], 8, f"S/ {float(p['precio_unitario']):.2f}", border=1, align='R', fill=fill)
            pdf.cell(widths[3], 8, f"S/ {float(p['subtotal']):.2f}", border=1, align='R', fill=fill)
            pdf.ln()
            fill = not fill
    else:
        pdf.cell(sum(widths), 8, "Aplica a toda la factura", border=1, align='C')

    pdf.set_font('Arial', 'B', 10)
    pdf.set_fill_color(0, 51, 102)
    pdf.set_text_color(255, 255, 255)

    pdf.cell(sum(widths[:3]), 8, "Subtotal (sin IGV):", border=1, align='R', fill=True)
    pdf.cell(widths[3], 8, f"S/ {subtotal_sin_igv:.2f}", border=1, align='R', fill=True)
    pdf.ln()

    pdf.cell(sum(widths[:3]), 8, "IGV (18%):", border=1, align='R', fill=True)
    pdf.cell(widths[3], 8, f"S/ {igv:.2f}", border=1, align='R', fill=True)
    pdf.ln()

    pdf.cell(sum(widths[:3]), 8, "TOTAL:", border=1, align='R', fill=True)
    pdf.cell(widths[3], 8, f"S/ {total:.2f}", border=1, align='R', fill=True)

    pdf.ln(15)
    pdf.set_font('Arial', 'I', 8)
    pdf.set_text_color(100, 100, 100)
    pdf.cell(0, 5, "Este documento es una representación impresa de una Nota Electrónica.", ln=1, align='C')

    pdf_bytes = pdf.output(dest='S').encode('latin1')
    pdf_output = BytesIO(pdf_bytes)
    pdf_output.seek(0)

    return send_file(pdf_output, mimetype='application/pdf')


@facturas_bp.route('/notas')
def listar_notas():
    conn = get_db_connection()
    cursor = conn.cursor()  # o DictCursor si usas PyMySQL

    cursor.execute("""
        SELECT 
            n.id,
            n.tipo_nota,
            n.motivo,
            n.monto,
            n.fecha_emision,
            f.numero_factura,
            c.nombre AS cliente_nombre,
            c.documento_tipo,
            c.documento_numero
        FROM notas_credito_debito n
        JOIN facturas f ON n.factura_id = f.id
        JOIN clientes c ON f.cliente_documento_numero = c.documento_numero
        ORDER BY n.fecha_emision DESC
    """)

    rows = cursor.fetchall()
    conn.close()

    notas = []
    for row in rows:
        notas.append({
            "nota": {
                "id": row["id"],
                "tipo_nota": row["tipo_nota"],
                "motivo": row["motivo"],
                "monto": row["monto"],
                "fecha_emision": row["fecha_emision"],
                "numero_factura": row["numero_factura"],
                "cliente_nombre": row["cliente_nombre"],
                "tipo_documento": row["documento_tipo"],  # Usamos el nombre correcto
                "numero_documento": row["documento_numero"]
            }
        })

    return render_template("facturas/listado_notas.html", notas=notas)


@facturas_bp.route('/imprimir-nota/<int:nota_id>')
def imprimir_nota(nota_id):
    conn = get_db_connection()
    cursor = conn.cursor()

    # Obtener cabecera de la nota + factura + cliente
    cursor.execute("""
        SELECT n.*, f.numero_factura, c.nombre AS cliente
        FROM notas_credito_debito n
        JOIN facturas f ON n.factura_id = f.id
        JOIN clientes c ON f.cliente_documento_numero = c.documento_numero
        WHERE n.id = %s
    """, (nota_id,))
    nota = cursor.fetchone()

    if not nota:
        cursor.close()
        conn.close()
        flash("Nota no encontrada", "danger")
        return redirect(url_for('facturas.listar_notas'))

    # Obtener detalle de productos afectados
    cursor.execute("""
        SELECT dn.producto_id, p.nombre AS producto_nombre,
               dn.cantidad, dn.precio_unitario, dn.subtotal
        FROM detalle_nota dn
        JOIN productos p ON dn.producto_id = p.id
        WHERE dn.nota_id = %s
    """, (nota_id,))
    detalle = cursor.fetchall()

    cursor.close()
    conn.close()

    return render_template('facturas/imprimir_nota.html', nota=nota, detalle=detalle)


@facturas_bp.route('/exportar-nota/<int:nota_id>')
def exportar_nota(nota_id):
    # Aquí podrías usar pdfkit, WeasyPrint o openpyxl según lo que prefieras
    # Por ahora retornamos un mensaje temporal
    return f"Exportar Nota {nota_id} a PDF (TODO)"


# Exportar a PDF
@facturas_bp.route('/exportar_facturas_pdf')
def exportar_facturas_pdf():
    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("""
    SELECT f.id, f.numero_factura, f.tipo_documento, c.nombre AS cliente, f.fecha_emision, f.total
    FROM facturas f
    JOIN clientes c ON f.cliente_documento_numero = c.documento_numero
    ORDER BY f.fecha_emision DESC
""")
    facturas = cursor.fetchall()

    cursor.close()
    conn.close()

    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 14)
    pdf.cell(0, 10, "Listado de Facturas", ln=True, align="C")

    pdf.set_font("Arial", "B", 12)
    # Definimos ancho de columnas
    ancho_cols = [10, 40, 60, 40, 30]
    encabezados = ["ID", "Número", "Cliente", "Fecha Emisión", "Total"]

    for i, encabezado in enumerate(encabezados):
        pdf.cell(ancho_cols[i], 10, encabezado, border=1, align="C")
    pdf.ln()

    pdf.set_font("Arial", "", 11)
    for f in facturas:
        pdf.cell(ancho_cols[0], 10, str(f['id']), border=1)
        pdf.cell(ancho_cols[1], 10, f['numero_factura'], border=1)
        pdf.cell(ancho_cols[2], 10, f['cliente'], border=1)
        # Si fecha_emision es datetime, formateamos, si es string, lo dejamos
        fecha = f['fecha_emision']
        if hasattr(fecha, "strftime"):
            fecha_str = fecha.strftime('%Y-%m-%d')
        else:
            fecha_str = str(fecha)
        pdf.cell(ancho_cols[3], 10, fecha_str, border=1)
        pdf.cell(ancho_cols[4], 10, f"{f['total']:.2f}", border=1, align='R')
        pdf.ln()

    # Generar PDF en memoria
    pdf_bytes = pdf.output(dest='S').encode('latin1')
    pdf_output = BytesIO(pdf_bytes)
    pdf_output.seek(0)

    return send_file(pdf_output, mimetype='application/pdf', as_attachment=True, download_name='facturas.pdf')

# Exportar a Excel
@facturas_bp.route('/exportar/excel')
def exportar_facturas_excel():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
    SELECT f.numero_factura, f.tipo_documento, c.nombre AS cliente, f.fecha_emision, f.total
    FROM facturas f
    JOIN clientes c ON f.cliente_documento_numero = c.documento_numero
    ORDER BY f.fecha_emision DESC
""")
    facturas = cursor.fetchall()
    cursor.close()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Facturas"

    columnas = ['Número', 'Cliente', 'Fecha emisión', 'Total']
    ws.append(columnas)

    for f in facturas:
        ws.append([
            f['numero_factura'],
            f['cliente'],
            f['fecha_emision'].strftime('%Y-%m-%d'),
            float(f['total'])
        ])

    # Ajustar ancho columnas
    for i, col in enumerate(columnas, 1):
        max_length = max(
            len(str(row[i-1])) if row[i-1] is not None else 0
            for row in ws.iter_rows(values_only=True)
        )
        ws.column_dimensions[get_column_letter(i)].width = max_length + 2

    excel_output = BytesIO()
    wb.save(excel_output)
    excel_output.seek(0)

    return send_file(excel_output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name='facturas.xlsx')


@facturas_bp.route('/api/next-number')
def get_next_number():
    tipo = request.args.get('tipo', 'factura')
    
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Obtener el último número según el tipo
        cursor.execute("""
            SELECT numero_factura 
            FROM facturas 
            WHERE tipo_documento = %s 
            ORDER BY id DESC 
            LIMIT 1
        """, (tipo,))
        
        ultimo_numero = cursor.fetchone()
        
        # Generar nuevo número
        if tipo == 'factura':
            prefijo = 'F'
        else:
            prefijo = 'B'
            
        if not ultimo_numero:
            nuevo_numero = f"{prefijo}00001"
        else:
            # Extraer el número secuencial (ej: F00001 -> 1)
            ultimo_num = int(ultimo_numero['numero_factura'][1:])
            nuevo_num = ultimo_num + 1
            nuevo_numero = f"{prefijo}{nuevo_num:05d}"
            
        return jsonify({"numero": nuevo_numero})
        
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()


#factura pdf 

@facturas_bp.route('/generar-pdf-preview', methods=['POST'])
def generar_pdf_preview():
    from fpdf import FPDF
    from io import BytesIO
    from datetime import datetime
    import json
    import os

    try:
        data = request.get_json()

        # Validación básica
        if not data or 'tipoDocumento' not in data or 'productos' not in data:
            return jsonify({"error": "Datos incompletos para generar el PDF"}), 400

        empresa = {
            'nombre': 'DULCE MAYOR S.A.C',
            'ruc': '10404129207',
            'direccion': 'CAR.CENTRAL KM. 6.5 INT. 258 OTR. CENTRO COMERCIAL LAS BRISAS DE ATE LIMA - LIMA - ATE',
            'telefono': '(01) 123-4567',
            'email': 'contacto@gmail.com',
            'logo': os.path.join(os.path.dirname(__file__), '..', 'static', 'img', 'logo.png')
        }

        pdf = FPDF(orientation='P', unit='mm', format='A4')
        pdf.add_page()
        pdf.set_margins(left=15, top=15, right=15)
        pdf.set_auto_page_break(auto=True, margin=15)

        if os.path.exists(empresa['logo']):
            pdf.image(empresa['logo'], x=15, y=10, w=30)
            pdf.set_y(40)
        else:
            pdf.set_y(20)

        pdf.set_font('Arial', 'B', 14)
        pdf.set_text_color(0, 0, 0)
        pdf.cell(0, 8, empresa['nombre'], ln=1)
        pdf.set_font('Arial', '', 10)
        pdf.cell(0, 5, f"RUC: {empresa['ruc']}", ln=1)

        direccion = empresa['direccion']
        max_width = 180
        if pdf.get_string_width(direccion) > max_width:
            words = direccion.split()
            current_line = []
            lines = []
            for word in words:
                test_line = ' '.join(current_line + [word])
                if pdf.get_string_width(test_line) < max_width:
                    current_line.append(word)
                else:
                    lines.append(' '.join(current_line))
                    current_line = [word]
            lines.append(' '.join(current_line))
            for line in lines:
                pdf.cell(0, 5, line, ln=1)
        else:
            pdf.cell(0, 5, direccion, ln=1)

        pdf.cell(0, 5, f"Tel: {empresa['telefono']} | Email: {empresa['email']}", ln=1)

        pdf.set_draw_color(200, 200, 200)
        pdf.line(15, pdf.get_y() + 5, 195, pdf.get_y() + 5)
        pdf.ln(10)

        pdf.set_font('Arial', 'B', 16)
        pdf.set_text_color(0, 51, 102)
        tipo_doc = "FACTURA" if data['tipoDocumento'] == 'factura' else "BOLETA"
        pdf.cell(0, 10, f"{tipo_doc} ELECTRÓNICA", ln=1, align='C')
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, f"N° {data['numeroDocumento']}", ln=1, align='C')

        pdf.set_font('Arial', '', 10)
        pdf.set_text_color(0, 0, 0)
        pdf.cell(0, 5, f"Fecha: {data['fecha']}", ln=1)

        pdf.ln(8)
        pdf.set_font('Arial', 'B', 12)
        pdf.cell(0, 8, "DATOS DEL CLIENTE", ln=1)

        pdf.set_fill_color(240, 240, 240)
        y_start = pdf.get_y()
        pdf.rect(15, y_start, 180, 25, style='F')

        pdf.set_font('Arial', '', 10)
        pdf.set_xy(20, y_start + 5)
        pdf.cell(0, 5, f"Nombre: {data['cliente'].get('nombre', '')}", ln=1)
        pdf.set_x(20)
        pdf.cell(0, 5, f"Documento: {data['cliente'].get('documento', '')}", ln=1)
        pdf.set_x(20)
        pdf.cell(0, 5, f"Dirección: {data['cliente'].get('direccion', '')}", ln=1)
        pdf.set_x(20)
        pdf.cell(0, 5, f"Teléfono: {data['cliente'].get('telefono', '')}", ln=1)

        pdf.set_y(pdf.get_y() + 5)

        pdf.ln(8)
        pdf.set_font('Arial', 'B', 12)
        pdf.cell(0, 8, "DETALLE DE PRODUCTOS", ln=1)

        pdf.set_fill_color(0, 51, 102)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font('Arial', 'B', 10)

        widths = [30, 70, 20, 30, 30]
        headers = ["Código", "Descripción", "Cant.", "P. Unitario", "Subtotal"]

        for i, header in enumerate(headers):
            pdf.cell(widths[i], 8, header, border=1, align='C', fill=True)
        pdf.ln()

        pdf.set_font('Arial', '', 10)
        pdf.set_text_color(0, 0, 0)
        fill = False

        for producto in data['productos']:
            pdf.set_fill_color(240, 240, 240) if fill else pdf.set_fill_color(255, 255, 255)

            pdf.cell(widths[0], 8, str(producto.get('codigo', '')), border=1, fill=fill)
            pdf.cell(widths[1], 8, str(producto.get('nombre', '')), border=1, fill=fill)
            pdf.cell(widths[2], 8, str(producto.get('cantidad', '')), border=1, align='C', fill=fill)
            pdf.cell(widths[3], 8, f"{float(producto.get('precio', 0)) :,.2f}", border=1, align='R', fill=fill)
            pdf.cell(widths[4], 8, f"{float(producto.get('subtotal', 0)) :,.2f}", border=1, align='R', fill=fill)
            pdf.ln()
            fill = not fill

        pdf.set_font('Arial', 'B', 10)
        pdf.set_fill_color(220, 220, 220)

        pdf.cell(sum(widths[:-1]), 8, "Subtotal:", border=1, align='R', fill=True)
        pdf.cell(widths[-1], 8, f"{float(data.get('subtotal', 0)) :,.2f}", border=1, align='R', fill=True)
        pdf.ln()

        pdf.cell(sum(widths[:-1]), 8, "IGV (18%):", border=1, align='R', fill=True)
        pdf.cell(widths[-1], 8, f"{float(data.get('igv', 0)) :,.2f}", border=1, align='R', fill=True)
        pdf.ln()

        pdf.set_fill_color(0, 51, 102)
        pdf.set_text_color(255, 255, 255)
        pdf.cell(sum(widths[:-1]), 8, "TOTAL:", border=1, align='R', fill=True)
        pdf.cell(widths[-1], 8, f"{float(data.get('total', 0)) :,.2f}", border=1, align='R', fill=True)

        pdf.ln(5)
        pdf.set_font('Arial', '', 10)
        pdf.set_text_color(0, 0, 0)
        pdf.cell(2, 8, f"Son: {data.get('totalEnLetras', '')}", ln=1)

        pdf.ln(15)
        pdf.set_font('Arial', 'I', 8)
        pdf.set_text_color(100, 100, 100)
        pdf.cell(0, 5, "¡Gracias por su compra!", ln=1, align='C')
        pdf.cell(0, 5, "Este documento es una representación impresa de un comprobante electrónico", ln=1, align='C')

        pdf_bytes = pdf.output(dest='S').encode('latin1')
        pdf_output = BytesIO(pdf_bytes)
        pdf_output.seek(0)

        return send_file(
            pdf_output,
            mimetype='application/pdf',
            as_attachment=False
        )

    except Exception as e:
        print(f"Error al generar PDF: {str(e)}")
        return jsonify({"error": str(e)}), 500


#integracion sunat:
# Token de acceso a la API de apisperu.com

@facturas_bp.route('/api/sunat', methods=['GET'])
def consultar_sunat():
    numero = request.args.get('numero', '').strip()
    if not numero:
        return jsonify({"error": "Documento inválido"}), 400

    if len(numero) == 8:
        url = f"https://dniruc.apisperu.com/api/v1/dni/{numero}?token={TOKEN}"
    elif len(numero) == 11:
        url = f"https://dniruc.apisperu.com/api/v1/ruc/{numero}?token={TOKEN}"
    else:
        return jsonify({"error": "Debe ser RUC (11) o DNI (8)"}), 400

    try:
        response = requests.get(url)
        if response.status_code == 200:
            data = response.json()
            if len(numero) == 8:
                nombre = f"{data.get('nombres', '')} {data.get('apellidoPaterno', '')} {data.get('apellidoMaterno', '')}"
                return jsonify({
                    "documento": numero,
                    "nombre": nombre,
                    "direccion": "-",
                    "telefono": "",
                    "estado": True
                })
            elif len(numero) == 11:
                return jsonify({
                    "documento": numero,
                    "nombre": data.get("razonSocial", ""),
                    "direccion": data.get("direccion", ""),
                    "telefono": "",
                    "estado": data.get("estado", "").lower() == "activo"
                })
        return jsonify({"error": "No se encontró información"}), 404
    except Exception as e:
        return jsonify({"error": "Error al consultar SUNAT", "detalles": str(e)}), 500

#registrar cliente buscado por sunat 
@facturas_bp.route('/api/registrar-cliente', methods=['POST'])
def registrar_cliente_api():
    data = request.get_json()
    if not data or 'documento' not in data:
        return jsonify({'error': 'Datos incompletos'}), 400

    conn = get_db_connection()
    cursor = conn.cursor()

    try:
        cursor.execute("""
            INSERT INTO clientes (documento_numero, nombre, direccion, telefono, estado)
            VALUES (%s, %s, %s, %s, 1)
        """, (
            data['documento'],
            data['nombre'],
            data['direccion'],
            data['telefono']
        ))
        conn.commit()
        return jsonify({'mensaje': 'Cliente registrado'}), 200
    except Exception as e:
        conn.rollback()
        return jsonify({'error': 'Error al guardar', 'detalle': str(e)}), 500
    finally:
        cursor.close()
        conn.close()

def numero_a_letras(numero):
    unidades = ['', 'UNO', 'DOS', 'TRES', 'CUATRO', 'CINCO', 'SEIS', 'SIETE', 'OCHO', 'NUEVE']
    decenas = ['DIEZ', 'ONCE', 'DOCE', 'TRECE', 'CATORCE', 'QUINCE', 'DIECISEIS', 'DIECISIETE', 'DIECIOCHO', 'DIECINUEVE']
    veintenas = ['VEINTE', 'VEINTIUNO', 'VEINTIDOS', 'VEINTITRES', 'VEINTICUATRO', 'VEINTICINCO', 
                 'VEINTISEIS', 'VEINTISIETE', 'VEINTIOCHO', 'VEINTINUEVE']
    decenas_completas = ['', '', 'TREINTA', 'CUARENTA', 'CINCUENTA', 'SESENTA', 'SETENTA', 'OCHENTA', 'NOVENTA']
    centenas = ['CIEN', 'CIENTO', 'DOSCIENTOS', 'TRESCIENTOS', 'CUATROCIENTOS', 'QUINIENTOS', 
                'SEISCIENTOS', 'SETECIENTOS', 'OCHOCIENTOS', 'NOVECIENTOS']
    
    entero = int(numero)
    decimal = int(round((numero - entero) * 100))
    
    if entero == 0:
        letras = 'CERO'
    elif entero < 10:
        letras = unidades[entero]
    elif entero < 20:
        letras = decenas[entero - 10]
    elif entero < 30:
        letras = veintenas[entero - 20]
    elif entero < 100:
        letras = decenas_completas[entero // 10]
        if entero % 10 != 0:
            letras += ' Y ' + unidades[entero % 10]
    elif entero < 1000:
        if entero == 100:
            letras = 'CIEN'
        else:
            letras = centenas[entero // 100]
            if entero % 100 != 0:
                letras += ' ' + numero_a_letras(entero % 100)
    else:
        letras = 'NÚMERO MUY GRANDE'
    
    return f"{letras} CON {decimal:02d}/100 SOLES"
